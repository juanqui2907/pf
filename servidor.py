"""
TerraShield — Servidor Flask
Soporta API de cálculo IEEE 80-2013, gestión de historiales y exportación.
"""
import os
import sys
import json
import datetime
import webbrowser
import threading

from flask import Flask, request, jsonify, send_from_directory

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def calcular_malla(p):
    import importlib, calculos_malla
    importlib.reload(calculos_malla)
    return calculos_malla.calcular_malla(p)

def listas_d(*args, **kwargs):
    import importlib, calculos_malla
    importlib.reload(calculos_malla)
    return calculos_malla.listas_d(*args, **kwargs)

PORT = 8000
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
HISTORIALES_DIR = os.path.join(BASE_DIR, 'historiales')

app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')

# ─── Crear carpeta historiales/ si no existe ──────────────────────────────────
os.makedirs(HISTORIALES_DIR, exist_ok=True)

# ─── Helpers de historial ─────────────────────────────────────────────────────

def _safe_nombre(nombre):
    """Nombre de archivo seguro: alfanumérico, guiones y guiones bajos."""
    safe = "".join(c for c in nombre if c.isalnum() or c in (' ', '-', '_')).strip()
    return safe.replace(' ', '_') or 'proyecto'

def _hist_path(nombre):
    return os.path.join(HISTORIALES_DIR, _safe_nombre(nombre) + '.json')

def _load_hist(nombre):
    path = _hist_path(nombre)
    if not os.path.exists(path):
        return []
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def _save_hist(nombre, hist):
    with open(_hist_path(nombre), 'w', encoding='utf-8') as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)

# ─── Archivos estáticos ───────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory(BASE_DIR, filename)

# ─── API: Gestión de proyectos ────────────────────────────────────────────────

@app.route('/proyectos', methods=['GET'])
def list_proyectos():
    """Lista los nombres de proyectos existentes (archivos .json en historiales/)."""
    files = [f[:-5] for f in os.listdir(HISTORIALES_DIR) if f.endswith('.json')]
    return jsonify(sorted(files))

@app.route('/proyectos/nuevo', methods=['POST'])
def nuevo_proyecto():
    """Crea el archivo de historial vacío para un proyecto si no existe."""
    data   = request.get_json(force=True, silent=True) or {}
    nombre = data.get('nombre', '').strip()
    if not nombre:
        return jsonify({'error': 'nombre requerido'}), 400
    path = _hist_path(nombre)
    if not os.path.exists(path):
        _save_hist(nombre, [])
    return jsonify({'ok': True, 'archivo': os.path.basename(path)})

@app.route('/proyectos/<nombre>', methods=['GET'])
def get_proyecto(nombre):
    """Devuelve el historial completo de un proyecto."""
    return jsonify(_load_hist(nombre))

# ─── API: /espaciamientos ─────────────────────────────────────────────────────

@app.route('/espaciamientos', methods=['GET'])
def get_espaciamientos():
    try:
        forma  = int(request.args.get('forma', 1))
        Lx     = float(request.args.get('Lx',   0))
        Ly     = float(request.args.get('Ly',   0))
        L      = float(request.args.get('L',    0))
        Lb     = float(request.args.get('Lb',   0))
        Li     = float(request.args.get('Li',   0))
        La     = float(request.args.get('La',   0))
        Lb_e   = float(request.args.get('Lb_e', 0))
        Lc_e   = float(request.args.get('Lc_e', 0))
    except (ValueError, TypeError) as e:
        return jsonify({"error": f"Parámetro inválido: {e}"}), 400

    resultado = listas_d(forma, Lx=Lx, Ly=Ly, L=L, Lb=Lb, Li=Li,
                         La=La, Lb_e=Lb_e, Lc_e=Lc_e)
    return jsonify(resultado)

# ─── API: /calcular ───────────────────────────────────────────────────────────

@app.route('/calcular', methods=['POST'])
def post_calcular():
    import traceback as _tb

    # ── DIAGNÓSTICO 1: body completo recibido ─────────────────────────────────
    data = request.get_json(force=True, silent=True)
    print("=" * 60)
    print("[DEBUG /calcular] Body recibido:")
    print(f"  keys       : {list(data.keys()) if data else 'None'}")
    print(f"  'proyecto' : {repr(data.get('proyecto')) if data else 'N/A'}")
    print("=" * 60)

    if not data:
        return jsonify({"error": "Body JSON requerido"}), 400

    required = ["forma", "rho", "h", "Lr", "b", "uv",
                "material_id", "I_falla", "tc", "tf", "XR", "Sf", "Tamb",
                "hs", "ts"]
    missing = [f for f in required if f not in data]
    if missing:
        return jsonify({"error": f"Faltan campos: {', '.join(missing)}"}), 400

    try:
        resultado = calcular_malla(data)
    except Exception as e:
        print("[ERROR /calcular]", _tb.format_exc())
        return jsonify({"error": str(e)}), 500

    # ── DIAGNÓSTICO 2: existencia de carpeta historiales/ ─────────────────────
    print(f"[DEBUG] HISTORIALES_DIR  : {HISTORIALES_DIR}")
    print(f"[DEBUG] carpeta existe   : {os.path.exists(HISTORIALES_DIR)}")
    print(f"[DEBUG] permisos escrit. : {os.access(HISTORIALES_DIR, os.W_OK)}")

    # Crear carpeta si falta (defensa extra)
    os.makedirs(HISTORIALES_DIR, exist_ok=True)

    # ── Guardar iteración en historial ────────────────────────────────────────
    proyecto = data.get('proyecto', '').strip()

    if not proyecto:
        print("[DEBUG] 'proyecto' vacío — historial NO guardado")
    else:
        hist_path = _hist_path(proyecto)
        print(f"[DEBUG] Ruta del JSON   : {hist_path}")
        print(f"[DEBUG] Archivo existe  : {os.path.exists(hist_path)}")

        try:
            # Crear archivo si no existe
            if not os.path.exists(hist_path):
                _save_hist(proyecto, [])
                print(f"[DEBUG] Archivo creado vacío: {hist_path}")

            hist = _load_hist(proyecto)
            print(f"[DEBUG] Iteraciones previas: {len(hist)}")

            sec3 = resultado.get('sec3', {})
            sec5 = resultado.get('sec5', {})
            sec7 = resultado.get('sec7', {})
            sec8 = resultado.get('sec8', {})
            sup  = resultado.get('superficie', {})

            Dx = float(data.get('Dx') or 0)
            Dy = float(data.get('Dy') or 0)
            D_otros = data.get('D') or data.get('Db') or data.get('D1') or data.get('D1_esc')
            if D_otros:
                D = float(D_otros)
            elif Dx > 0 and Dy > 0:
                D = round((Dx + Dy) / 2, 4)
            else:
                D = Dx or Dy

            iteracion = {
                'iter':      len(hist) + 1,
                'timestamp': datetime.datetime.now().isoformat(timespec='seconds'),
                'Lx':        data.get('Lx', data.get('L', 0)),
                'Ly':        data.get('Ly', data.get('L', 0)),
                'D':         round(float(D), 4) if D else 0,
                'h':         data.get('h', 0),
                'Lr':        data.get('Lr', 0),
                'Nr':        sec3.get('nR', 0),
                'hs':        data.get('hs', 0),
                'rho_s':     round(float(sup.get('rho_s', 0)), 2),
                'Rg':        round(float(sec5.get('Rg', 0)), 4),
                'GPR':       round(float(sec7.get('GPR_V', 0)), 2),
                'Em':        round(float(sec8.get('Em_V', 0)), 2),
                'Es':        round(float(sec8.get('Es_V', 0)), 2),
                'ok_Rg':     float(sec5.get('Rg', 999)) <= 1.0,
                'ok_GPR':    bool(sec7.get('gpr_ok', False)),
                'ok_Em':     bool(sec8.get('em_ok', False)),
                'ok_Es':     bool(sec8.get('es_ok', False)),
                'payload':   data,
            }

            # ── DIAGNÓSTICO 3: justo antes y después de escribir ──────────────
            print(f"[DEBUG] Guardando iteración #{iteracion['iter']}...")
            hist.append(iteracion)
            _save_hist(proyecto, hist)
            print(f"[DEBUG] Guardado exitoso — {hist_path}")

            # ── DIAGNÓSTICO 4: verificar que el archivo tiene contenido ───────
            size = os.path.getsize(hist_path)
            print(f"[DEBUG] Tamaño del archivo tras escritura: {size} bytes")

        except Exception as e:
            print(f"[ERROR] Fallo al guardar historial: {e}")
            print(_tb.format_exc())

    return jsonify(resultado)

# ─── API: Exportar Excel ──────────────────────────────────────────────────────

@app.route('/proyectos/<nombre>/exportar/excel', methods=['GET'])
def exportar_excel(nombre):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from flask import send_file
        import io
    except ImportError:
        return jsonify({'error': 'Instale openpyxl: pip install openpyxl'}), 500

    hist = _load_hist(nombre)
    aprobada = next(
        (it for it in reversed(hist)
         if it.get('ok_Rg') and it.get('ok_GPR') and it.get('ok_Em') and it.get('ok_Es')),
        None
    )
    if not aprobada:
        return jsonify({'error': 'No existe ninguna iteración con todos los criterios aprobados'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = 'Diseño Final'

    header_fill = PatternFill('solid', fgColor='0A3D91')
    header_font = Font(color='FFFFFF', bold=True)
    green_fill  = PatternFill('solid', fgColor='C6EFCE')
    center      = Alignment(horizontal='center')

    headers = ['#', 'Fecha', 'Lx (m)', 'Ly (m)', 'D (m)', 'h (m)',
               'Lr (m)', 'Nr', 'hs (m)', 'ρs (Ω·m)',
               'Rg (Ω)', 'GPR (V)', 'Em (V)', 'Es (V)',
               'ok Rg', 'ok GPR', 'ok Em', 'ok Es']
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    fila = [
        aprobada.get('iter'), aprobada.get('timestamp'),
        aprobada.get('Lx'), aprobada.get('Ly'), aprobada.get('D'),
        aprobada.get('h'), aprobada.get('Lr'), aprobada.get('Nr'),
        aprobada.get('hs'), round(aprobada.get('rho_s', 0), 2),
        aprobada.get('Rg'), aprobada.get('GPR'),
        aprobada.get('Em'), aprobada.get('Es'),
        '✔', '✔', '✔', '✔',
    ]
    ws.append(fila)
    for col in range(15, 19):
        ws.cell(row=2, column=col).fill = green_fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"TerraShield_{_safe_nombre(nombre)}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── API: Exportar PDF ────────────────────────────────────────────────────────

@app.route('/proyectos/<nombre>/exportar/pdf', methods=['GET'])
def exportar_pdf(nombre):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from flask import send_file
        import io
    except ImportError:
        return jsonify({'error': 'Instale reportlab: pip install reportlab'}), 500

    hist = _load_hist(nombre)
    aprobada = next(
        (it for it in reversed(hist)
         if it.get('ok_Rg') and it.get('ok_GPR') and it.get('ok_Em') and it.get('ok_Es')),
        None
    )
    if not aprobada:
        return jsonify({'error': 'No existe ninguna iteración con todos los criterios aprobados'}), 400

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=40, bottomMargin=40,
                            leftMargin=50, rightMargin=50)
    styles = getSampleStyleSheet()
    story  = []

    story.append(Paragraph('TerraShield — Malla de puesta a tierra IEEE 80-2013', styles['Title']))
    story.append(Paragraph(f'Proyecto: <b>{nombre}</b>', styles['Normal']))
    story.append(Paragraph(
        f'Fecha de exportación: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}',
        styles['Normal']
    ))
    story.append(Spacer(1, 20))
    story.append(Paragraph('Diseño Final Aprobado', styles['Heading2']))
    story.append(Spacer(1, 8))

    tabla_data = [
        ['Parámetro', 'Valor', 'Unidad'],
        ['Iteración N°',               str(aprobada.get('iter', '—')),              ''],
        ['Fecha de cálculo',           aprobada.get('timestamp', '—'),              ''],
        ['Lx — longitud X',            str(aprobada.get('Lx', '—')),               'm'],
        ['Ly — longitud Y',            str(aprobada.get('Ly', '—')),               'm'],
        ['D — espaciamiento',          str(aprobada.get('D', '—')),                'm'],
        ['h — profundidad de malla',   str(aprobada.get('h', '—')),                'm'],
        ['Lr — longitud de varilla',   str(aprobada.get('Lr', '—')),               'm'],
        ['Nr — número de varillas',    str(aprobada.get('Nr', '—')),               ''],
        ['hs — capa superficial',      str(aprobada.get('hs', '—')),               'm'],
        ['ρs — resistividad sup.',     str(round(aprobada.get('rho_s', 0), 2)),    'Ω·m'],
        ['Rg — resist. de malla',      str(aprobada.get('Rg', '—')),               'Ω  ✔'],
        ['GPR — potencial de tierra',  str(aprobada.get('GPR', '—')),              'V  ✔'],
        ['Em — voltaje de malla',      str(aprobada.get('Em', '—')),               'V  ✔'],
        ['Es — voltaje de paso',       str(aprobada.get('Es', '—')),               'V  ✔'],
    ]

    t = Table(tabla_data, colWidths=[210, 110, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#0A3D91')),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0, 1),  (-1, -1), [colors.white, colors.HexColor('#EEF3FF')]),
        ('GRID',          (0, 0),  (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('FONTSIZE',      (0, 0),  (-1, -1), 9),
        ('TOPPADDING',    (0, 0),  (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0),  (-1, -1), 5),
        ('LEFTPADDING',   (0, 0),  (-1, -1), 8),
    ]))
    story.append(t)

    doc.build(story)
    output.seek(0)

    fname = f"TerraShield_{_safe_nombre(nombre)}.pdf"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')

# ─── Arranque ─────────────────────────────────────────────────────────────────

def abrir_navegador():
    import time
    time.sleep(1.2)
    webbrowser.open(f'http://localhost:{PORT}')

if __name__ == '__main__':
    threading.Thread(target=abrir_navegador, daemon=True).start()
    print("=" * 46)
    print("  TERRASHIELD v1.0.0")
    print(f"  Servidor Flask en http://localhost:{PORT}")
    print("  Cierre esta ventana para detener.")
    print("=" * 46)
    app.run(host='0.0.0.0', port=PORT, debug=False)
